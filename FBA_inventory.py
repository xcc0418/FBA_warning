# import openpyxl
import openpyxl
import requests
import pymysql
import json
import time
import datetime
import email.mime.multipart
import smtplib
from email.mime.text import MIMEText
import os
from email.mime.application import MIMEApplication
import zipfile
# import global_var


class Find_order():
    def __init__(self):
        self.s = requests.Session()
        login_url = 'https://erp.lingxing.com/api/passport/login'
        # 请求头
        headers = {'Host': 'erp.lingxing.com',
                   'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                    , 'Referer': 'https://erp.lingxing.com/login',
                   'Accept': 'application/json, text/plain, */*',
                   'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                   'Accept-Encoding': 'gzip, deflate, br',
                   'Content-Type': 'application/json;charset=utf-8',
                   'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                   'X-AK-Company-Id': '90136229927150080',
                   'X-AK-Request-Source': 'erp',
                   'X-AK-ENV-KEY': 'SAAS-10',
                   'X-AK-Version': '1.0.0.0.0.023',
                   'X-AK-Zid': '109810',
                   'Content-Length': '114',
                   'Origin': 'https://erp.lingxing.com',
                   'Connection': 'keep-alive'}
        # 传递用户名和密码
        data = {'account': 'IT-Test', 'pwd': 'IT-Test'}
        data = json.dumps(data)
        self.s.post(login_url, headers=headers, data=data)
        self.auth_token = None
        # self.write_sql()

    def sql(self):
        self.connection = pymysql.connect (host='3354n8l084.goho.co',  # 数据库地址
                                           port=24824,
                                           user='test_user',  # 数据库用户名
                                           password='a123456',  # 数据库密码
                                           db='storage',  # 数据库名称
                                           charset='utf8',
                                           cursorclass=pymysql.cursors.DictCursor)
        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.connection.cursor()

    def sql_close(self):
        self.cursor.close()
        self.connection.close()

    def check(self):
        self.sql()
        sql = "SELECT `flag_num` FROM `flag`.`amazon_form_flag` WHERE `flag_name`='FBA_warning'"
        self.cursor.execute(sql)
        result = self.cursor.fetchone()
        self.sql_close()
        if int(result['flag_num']) == 1:
            return False
        else:
            return True

    def check_1(self):
        self.sql()
        sql = "UPDATE `flag`.`amazon_form_flag` SET `flag_num`=1 WHERE `flag_name`='FBA_warning'"
        self.cursor.execute(sql)
        self.connection.commit()
        self.sql_close()

    def check_0(self):
        self.sql()
        sql = "UPDATE `flag`.`amazon_form_flag` SET `flag_num`=0 WHERE `flag_name`='FBA_warning'"
        self.cursor.execute(sql)
        self.connection.commit()
        self.sql_close()

    def get_FBA(self):
        self.makdir()
        get_url = "https://erp.lingxing.com/api/storage/fbaExport?status=&cid=&bid=&attribute=&asin_principal=&" \
                  "search_field=sku&offset=0&length=200&fulfillment_channel_type=&is_hide_zero_stock=0&" \
                  "is_parant_asin_merge=0&is_inclide_header=0&is_contain_del_ls=0&is_cost_page=0&" \
                  "req_time_sequence=/api/storage/fbaExport$$"
        get_headers = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/erp/msupply/fbaInventory'}
        get_msg = self.s.get(get_url, headers=get_headers)
        get_msg = json.loads(get_msg.text)
        report_id = ''
        print(get_msg)
        if get_msg['code'] == 1 and get_msg['msg'] == 'success':
            report_id = get_msg['report_id']
        if report_id:
            time.sleep(300)
            file_download_url = f"https://erp.lingxing.com/api/download/downloadCenterReport/downloadResource?report_id={report_id}"
            # print(file_download_url)
            get_headers = {'user-agent': 'Mozilla/5.0', 'Referer': 'https://erp.lingxing.com/replenishmentAdvice'}
            download_file = self.s.get(file_download_url, headers=get_headers, stream=False)
            with open('D:/FBA库存提醒/FBA_inventory.zip', 'wb') as q:
                q.write(download_file.content)
            return report_id

    def write_sql(self):
        wb = openpyxl.load_workbook('D:/FBA库存提醒/FBA_inventory.xlsx')
        wb_sheet = wb.active
        row1 = wb_sheet.max_row
        for i in range(row1, 0, -1):
            cell_value1 = wb_sheet.cell(row=i, column=1).value
            if cell_value1:
                row1 = i
                break
        self.sql()
        for i in range(2, row1+1):
            f_asin = wb_sheet.cell(row=i, column=1).value
            sku = wb_sheet.cell(row=i, column=7).value
            fnsku = wb_sheet.cell(row=i, column=6).value
            msku = wb_sheet.cell(row=i, column=5).value
            if sku:
                country = self.find_country(msku, f_asin)
                print(country)
                sql = "update `process`.`fba_inventory` set `国家` = '%s' where `SKU` = '%s' and `FNSKU` = " \
                      "'%s' and `父ASIN` = '%s'" % (country, sku, fnsku, f_asin)
                self.cursor.execute(sql)
        self.connection.commit()
        self.sql_close()

    def clear_sql(self):
        # report_id = self.get_FBA()
        # file = zipfile.ZipFile('D:/FBA库存提醒/FBA_inventory.zip')
        # file.extractall('D:/FBA库存提醒/')
        # file.close()
        report_id = '481288679238819841'
        filename = f"D:/FBA库存提醒/仓库管理-FBA库存-{report_id}.xlsx"
        wb = openpyxl.load_workbook(filename)
        wb_sheet = wb.active
        row1 = wb_sheet.max_row
        column1 = wb_sheet.max_column
        for i in range(row1, 0, -1):
            cell_value1 = wb_sheet.cell(row=i, column=1).value
            if cell_value1:
                row1 = i
                break
        list_wb_heard = []
        for i in range(1, column1 + 1):
            heard = wb_sheet.cell(row=1, column=i).value
            if heard:
                list_wb_heard.append(heard)
        print(row1)
        self.sql()
        for i in range(2, row1 + 1):
            f_asin = wb_sheet.cell(row=i, column=(list_wb_heard.index('父ASIN') + 1)).value
            asin = wb_sheet.cell(row=i, column=(list_wb_heard.index('ASIN') + 1)).value
            sku = wb_sheet.cell(row=i, column=(list_wb_heard.index('SKU') + 1)).value
            fnsku = wb_sheet.cell(row=i, column=(list_wb_heard.index('FNSKU') + 1)).value
            name = wb_sheet.cell(row=i, column=(list_wb_heard.index('品名') + 1)).value
            num = int(wb_sheet.cell(row=i, column=(list_wb_heard.index('FBA可售') + 1)).value)
            # msku = wb_sheet.cell(row=i, column=(list_wb_heard.index('MSKU') + 1)).value
            # country = self.find_country(msku, f_asin)
            country = wb_sheet.cell(row=i, column=(list_wb_heard.index('所属仓库') + 1)).value
            sql = "insert into `process`.`fba_inventory`(`父ASIN`, `ASIN`, `SKU`, `FNSKU`,`国家`, `品名`, " \
                  "`FBA库存_旧`, `FBA库存_新`)values('%s', '%s', '%s', '%s', '%s', '%s', %d, %d)" \
                  "" % (f_asin, asin, sku, fnsku, country, name, num, num)
            self.cursor.execute(sql)
            if i % 100 == 0 or i == row1:
                print(i)
                self.connection.commit()
        self.sql_close()

    def change_sql(self):
        report_id = '481288679238819841'
        filename = f"D:/FBA库存提醒/仓库管理-FBA库存-{report_id}.xlsx"
        wb = openpyxl.load_workbook (filename)
        wb_sheet = wb.active
        row1 = wb_sheet.max_row
        column1 = wb_sheet.max_column
        for i in range (row1, 0, -1):
            cell_value1 = wb_sheet.cell (row=i, column=1).value
            if cell_value1:
                row1 = i
                break
        list_wb_heard = []
        for i in range (1, column1 + 1):
            heard = wb_sheet.cell (row=1, column=i).value
            if heard:
                list_wb_heard.append (heard)
        print (row1)
        self.sql ()
        for i in range (2, row1 + 1):
            country = wb_sheet.cell (row=i, column=(list_wb_heard.index ('所属仓库') + 1)).value

    def read_excl(self, report_id):
        # file = zipfile.ZipFile('D:/FBA库存提醒/FBA_inventory.zip')
        # file.extractall('D:/FBA库存提醒/')
        # file.close()
        filename = f"D:/FBA库存提醒/仓库管理-FBA库存-{report_id}.xlsx"
        wb = openpyxl.load_workbook(filename)
        wb_sheet = wb.active
        row1 = wb_sheet.max_row
        column1 = wb_sheet.max_column
        for i in range(row1, 0, -1):
            cell_value1 = wb_sheet.cell(row=i, column=1).value
            if cell_value1:
                row1 = i
                break
        print(row1)
        list_wb_heard = []
        for i in range(1, column1 + 1):
            heard = wb_sheet.cell(row=1, column=i).value
            if heard:
                list_wb_heard.append(heard)
        self.sql()
        number = 0
        for i in range(2, row1 + 1):
            if i == row1:
                print('个数：', number)
                self.connection.commit()
            f_asin = wb_sheet.cell(row=i, column=(list_wb_heard.index('父ASIN') + 1)).value
            asin = wb_sheet.cell(row=i, column=(list_wb_heard.index('ASIN') + 1)).value
            sku = wb_sheet.cell(row=i, column=(list_wb_heard.index('SKU') + 1)).value
            fnsku = wb_sheet.cell(row=i, column=(list_wb_heard.index('FNSKU') + 1)).value
            name = wb_sheet.cell(row=i, column=(list_wb_heard.index('品名') + 1)).value
            num = int(wb_sheet.cell(row=i, column=(list_wb_heard.index('FBA可售') + 1)).value)
            # msku = wb_sheet.cell(row=i, column=(list_wb_heard.index('MSKU') + 1)).value
            # country = self.find_country(msku, f_asin)
            country = wb_sheet.cell(row=i, column=(list_wb_heard.index('所属仓库') + 1)).value
            sql1 = "select * from `process`.`fba_inventory` where `FNSKU` = '%s' and `父ASIN` = '%s' and `国家` = '%s'" % (fnsku, f_asin, country)
            self.cursor.execute(sql1)
            result = self.cursor.fetchone()
            # sql = ''
            if result:
                # print(sku)
                if result['FBA库存_新'] == num:
                    continue
                if result['SKU'] == 'None':
                    print(fnsku)
                    sql = "update `process`.`fba_inventory` set `FBA库存_新` = %d, `FBA库存_旧` = %d , `SKU` = '%s', `品名` = '%s' where `国家` = '%s' and `FNSKU` = " \
                          "'%s' and `父ASIN` = '%s'" % (num, result['FBA库存_新'], sku, name, country, fnsku, f_asin)
                else:
                    sql = "update `process`.`fba_inventory` set `FBA库存_新` = %d, `FBA库存_旧` = %d where `FNSKU` = " \
                          "'%s' and `父ASIN` = '%s' and `国家` = '%s'" % (num, result['FBA库存_新'], fnsku, f_asin, country)
                # print(sql)
            else:
                sql = "insert into `process`.`fba_inventory`(`父ASIN`, `ASIN`, `SKU`, `FNSKU`,`国家`, `品名`, " \
                      "`FBA库存_旧`, `FBA库存_新`)values('%s', '%s', '%s', '%s', '%s', '%s', %d, %d)" \
                      "" % (f_asin, asin, sku, fnsku, country, name, num, num)
            # print(sql)
            if sql:
                self.cursor.execute(sql)
                number += 1
            if number % 100 == 0 or i == row1:
                print(i)
                print('个数：', number)
                self.connection.commit()
                self.sql_close()
                self.sql()
        self.sql_close()

    def get_fnsku(self):
        self.sql()
        list_fnsku = []
        time_now = datetime.datetime.now().strftime("%Y%m%d%H%M")
        sql = "select * from `process`.`fba_inventory` where `FBA库存_新` = 0"
        self.cursor.execute(sql)
        result = self.cursor.fetchall()
        for i in result:
            num_old = int(i['FBA库存_旧'])
            num_new = int(i['FBA库存_新'])
            time_get = str(i['修改时间'])[0:10]
            time_flag = str(datetime.datetime.now().strftime("%Y-%m-%d"))
            if num_new == 0 and num_old > 0 and time_flag == time_get:
                # country = self.find_country(i['FNSKU'])
                list_fnsku.append([i['父ASIN'], i['ASIN'], i['SKU'], i['FNSKU'], i['国家'], i['品名'], i['FBA库存_新']])
                sql = "insert into `process`.`FBA_warning`(`父_ASIN`, `SKU`, `FNSKU`, `品名`, `国家`, `状态`)values" \
                      "('%s', '%s', '%s', '%s', '%s', '未处理')" % (i['父ASIN'], i['SKU'], i['FNSKU'], i['品名'], i['国家'])
                self.cursor.execute(sql)
                self.connection.commit()
        print(list_fnsku)
        if list_fnsku:
            wb = openpyxl.Workbook()
            wb_sheet = wb.active
            wb_sheet.append(['父ASIN', 'ASIN', 'SKU', 'FNSKU', '国家', '品名', 'FBA库存_新'])
            for i in list_fnsku:
                wb_sheet.append(i)
            filename = f"D:/FBA库存提醒/FBA库存提醒_{time_now}.xlsx"
            wb.save(filename)
            time_flag = int(datetime.datetime.now().strftime("%H"))
            if 2 < time_flag < 6:
                self.smtplib_email(filename)
        self.sql_close()

    def makdir(self):
        folder = os.path.exists("D:/FBA库存提醒")
        if not folder:
            os.makedirs("D:/FBA库存提醒")

    def smtplib_email(self, filename):
        # for i in list_sku:
        #     sku = i*/
        time_now = time.strftime("%Y-%m-%d", time.localtime())
        my_send = 'xiechangcong@getoo.store'
        my_password = 'XCc200418'
        receivers = 'wenzhichao@getoo.store'  # 接收邮件，可设置为你的QQ邮箱或者其他邮箱 nixuemin@getoo.store  2371138547@qq.com
        ret = True
        # 三个参数：第一个为文本内容，第二个 plain 设置文本格式，第三个 utf-8 设置编码
        message = email.mime.multipart.MIMEMultipart()
        message['from'] = my_send
        message['to'] = receivers
        subject = 'FBA库存提醒'
        message['Subject'] = subject
        content = "FBA库存提醒："
        text = email.mime.text.MIMEText(content, 'plain', 'utf-8')
        message.attach(text)
        att = MIMEApplication(open(f"{filename}", 'rb').read())
        att.add_header('Content-Disposition', 'attachment', filename=f"FBA库存提醒-{time_now}.xlsx")
        message.attach(att)
        try:
            # message['From'] = formataddr(["XCC", my_send])  # 发送者https://union.jd.com/proManager/index?categories=16750,16755&pageNo=1
            # message['To'] = formataddr(["测试", receivers])  # 接收者
            server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)
            server.login(my_send, my_password)
            server.sendmail(my_send, [receivers, ], message.as_string())
            server.quit()
        except Exception as e:
            print(e)
            ret = False
        if ret:
            print("邮件发送成功")
        else:
            self.smtplib_error(e)
            print("邮件发送失败")

    def smtplib_error(self, e):
        time_now = time.strftime("%Y-%m-%d", time.localtime())
        my_send = 'xiechangcong@getoo.store'
        my_password = 'XCc200418'
        receivers = '2371138547@qq.com'  # 接收邮件，可设置为你的QQ邮箱或者其他邮箱 nixuemin@getoo.store  2371138547@qq.com
        ret = True
        # 三个参数：第一个为文本内容，第二个 plain 设置文本格式，第三个 utf-8 设置编码
        message = email.mime.multipart.MIMEMultipart()
        message['from'] = my_send
        message['to'] = receivers
        subject = 'FBA可售预警'
        message['Subject'] = subject
        content = f"{e}"
        text = email.mime.text.MIMEText(content, 'plain', 'utf-8')
        message.attach(text)
        # att = MIMEApplication (open (f"{filename}", 'rb').read ())
        # att.add_header ('Content-Disposition', 'attachment', filename=f"常用物料采购-{time_now}.xlsx")
        # message.attach (att)
        try:
            # message['From'] = formataddr(["XCC", my_send])  # 发送者
            # message['To'] = formataddr(["测试", receivers])  # 接收者
            server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)
            server.login(my_send, my_password)
            server.sendmail(my_send, [receivers, ], message.as_string())
            server.quit()
        except Exception as e:
            print(e)
            ret = False
        if ret:
            print("邮件发送成功")
        else:
            print("邮件发送失败")

    def find_country(self, msku, f_asin):
        auth_token = self.s.cookies.get('auth-token')
        # print(auth_token)
        auth_token = auth_token.replace('%25', '%')
        auth_token = auth_token.replace('%23', '#')
        auth_token = auth_token.replace('%26', '&')
        auth_token = auth_token.replace('%2B', '+')
        auth_token = auth_token.replace('%28', '(')
        auth_token = auth_token.replace('%29', ')')
        auth_token = auth_token.replace('%2F', '/')
        auth_token = auth_token.replace('%3D', '=')
        auth_token = auth_token.replace('%3F', '?')
        post_url = 'https://gw.lingxingerp.com/listing-api/api/product/showOnline'
        post_headers = {'Host': 'gw.lingxingerp.com',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:95.0) Gecko/20100101 Firefox/95.0'
                        , 'Referer': 'https://erp.lingxing.com/',
                        'auth-token': auth_token,
                        'Accept': 'application/json, text/plain, */*',
                        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Content-Type': 'application/json;charset=utf-8',
                        'X-AK-Request-Id': 'e7f7b81a-fafd-4031-8964-00376ae24d07',
                        'X-AK-Company-Id': '90136229927150080',
                        'X-AK-Request-Source': 'erp',
                        'X-AK-ENV-KEY': 'SAAS-10',
                        'X-AK-Version': '1.0.0.0.0.023',
                        'X-AK-Zid': '109810',
                        'Content-Length': '909',
                        'Origin': 'https://erp.lingxing.com',
                        'Connection': 'keep-alive'}
        data = {"offset": 0, "length": 500, "search_field": "msku", "search_value": [f"{msku}"], "exact_search": 1,
                "sids": "", "status": "", "is_pair": "", "fulfillment_channel_type": "", "global_tag_ids": "",
                "req_time_sequence": "/listing-api/api/product/showOnline$$"}
        data = json.dumps(data)
        post_msg = self.s.post(post_url, headers=post_headers, data=data)
        post_msg = json.loads(post_msg.text)
        # print(post_msg)
        if post_msg['code'] == 1 and post_msg['msg'] == '成功':
            for i in post_msg['data']['list']:
                if i['msku'] == msku and i['parent_asin'] == f_asin:
                    return i['marketplace']


if __name__ == '__main__':
    order = Find_order()
    try:
        # order.clear_sql()
        # report_id = order.get_FBA()
        report_id = '482375840717668352'
        order.read_excl(report_id)
        # print("succeed2")
        order.get_fnsku()
    except Exception as e:
        print(e)
        order.smtplib_error(e)
